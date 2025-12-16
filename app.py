import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date
from openpyxl import load_workbook

st.set_page_config(page_title="Summary Billing Transport", layout="wide")

st.title("Summary Billing Transport")
st.caption("Converted from VBA → Python (Streamlit Cloud)")

uploaded_file = st.file_uploader(
    "Upload Excel file (Summary Billing To Customer...)",
    type=["xlsx", "xlsm"]
)

if not uploaded_file:
    st.stop()

# ===============================
# LOAD WORKBOOK
# ===============================
wb = load_workbook(uploaded_file, data_only=True)

sheet_names = wb.sheetnames

# ===============================
# VALIDATION
# ===============================
required_sheets = ["Main", "Cargo and Weight", "Sell Price"]
missing = [s for s in required_sheets if s not in sheet_names]

if missing:
    st.error(f"Missing required sheets: {missing}")
    st.stop()

# ===============================
# READ MAIN DATE (YEAR / MONTH)
# ===============================
main_ws = wb["Main"]
main_date = main_ws.cell(row=6, column=7).value

YEAR = main_date.year
MONTH = main_date.month

# ===============================
# LOAD SUPPORT TABLES
# ===============================
# ===============================
# LOAD SUPPORT TABLES (SAFE)
# ===============================

cargo_raw = pd.DataFrame(wb["Cargo and Weight"].values)
cargo_df = cargo_raw.iloc[1:, [1, 4]]   # B = Part Number, E = Weight
cargo_df.columns = ["Part Number", "Weight"]

sell_raw = pd.DataFrame(wb["Sell Price"].values)
sell_df = sell_raw.iloc[1:, [0, 2, 3, 4]]  # A, C, D, E
sell_df.columns = ["Post Code", "Area", "Min Charge", "Rate/KG"]

# ===============================
# COLLECT TEMP DATA
# ===============================
temp_rows = []

for sheet in sheet_names:
    if sheet.isnumeric():
        ws = wb[sheet]
        day = int(sheet)
        delivery_date = date(YEAR, MONTH, day)

        r = 9
        while ws.cell(r, 1).value:
            temp_rows.append({
                "Order Date": delivery_date,
                "DU": ws.cell(r, 1).value,
                "Order": ws.cell(r, 2).value,
                "DU-Order": ws.cell(r, 3).value,
                "CM Code": ws.cell(r, 4).value,
                "Sold To": ws.cell(r, 5).value,
                "Destination Code": ws.cell(r, 6).value,
                "Ship To": ws.cell(r, 7).value,
                "Address1": ws.cell(r, 8).value,
                "Address2": ws.cell(r, 9).value,
                "Province": ws.cell(r, 10).value,
                "Post Code": ws.cell(r, 11).value,
                "Part No": ws.cell(r, 13).value,
                "Pick Qty": ws.cell(r, 14).value,
                "Remark": ws.cell(r, 16).value
            })
            r += 1

temp_df = pd.DataFrame(temp_rows)
temp_df.sort_values(["Order Date", "DU-Order"], inplace=True)

# ===============================
# BUILD SUMMARY
# ===============================
summary_rows = []

for du_order, grp in temp_df.groupby("DU-Order", sort=False):
    header = grp.iloc[0]

    sum_qty = 0
    sum_weight = 0

    post_code = header["Post Code"]
    sell_match = sell_df[sell_df["Post Code"] == post_code]

    if sell_match.empty:
        area = "Post Code Not Found"
        rate = 0
        min_charge = 0
    else:
        area = sell_match.iloc[0]["Area"]
        rate = sell_match.iloc[0]["Rate/KG"]
        min_charge = sell_match.iloc[0]["Min Charge"]

    header_row_index = len(summary_rows)

    summary_rows.append({
        "Order Date": header["Order Date"],
        "DU": "",
        "DU-Order": header["DU-Order"],
        "CM Code.": header["CM Code"],
        "Sold To": header["Sold To"],
        "Destination Code": header["Destination Code"],
        "Ship To": header["Ship To"],
        "Address 1": header["Address1"],
        "Address 2": header["Address2"],
        "Province": header["Province"],
        "Post Code": header["Post Code"],
        "Area": area,
        "Total Pick Q'TY": "",
        "Ship Total WT": "",
        "Up 10KG/Chg": "",
        "Rate/KG": rate,
        "Min/Charge": min_charge,
        "All Charge": "",
        "Pick Date": "",
        "Transport": "",
        "Premium": "",
        "Remark": header["Remark"],
        "Part No.": "",
        "Pick Q'TY": "",
        "Total Weight": ""
    })

    for _, row in grp.iterrows():
        part = row["Part No"]
        qty = row["Pick Qty"]
        sum_qty += qty

        cargo_match = cargo_df[cargo_df["Part Number"] == part]
        if cargo_match.empty:
            weight = "Part Not Found"
        else:
            weight = qty * cargo_match.iloc[0]["Weight"]
            sum_weight += weight

        summary_rows.append({
            "Order Date": "",
            "DU": "",
            "DU-Order": "",
            "CM Code.": "",
            "Sold To": "",
            "Destination Code": "",
            "Ship To": "",
            "Address 1": "",
            "Address 2": "",
            "Province": "",
            "Post Code": "",
            "Area": "",
            "Total Pick Q'TY": "",
            "Ship Total WT": "",
            "Up 10KG/Chg": "",
            "Rate/KG": "",
            "Min/Charge": "",
            "All Charge": "",
            "Pick Date": "",
            "Transport": "",
            "Premium": "",
            "Remark": "",
            "Part No.": part,
            "Pick Q'TY": qty,
            "Total Weight": weight
        })

    up_10 = max(sum_weight - 10, 0)
    all_charge = (up_10 * rate) + min_charge
    transport = "STL" if area == "BKK" else "DASH"

    summary_rows[header_row_index]["Total Pick Q'TY"] = sum_qty
    summary_rows[header_row_index]["Ship Total WT"] = sum_weight
    summary_rows[header_row_index]["Up 10KG/Chg"] = up_10
    summary_rows[header_row_index]["All Charge"] = all_charge
    summary_rows[header_row_index]["Transport"] = transport

summary_df = pd.DataFrame(summary_rows)

# ===============================
# EXPORT TO EXCEL
# ===============================
output = BytesIO()

with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

output.seek(0)

# ===============================
# DOWNLOAD
# ===============================
st.success("Processing completed successfully")

st.download_button(
    "Download Summary File",
    data=output,
    file_name="Summary_Billing_Output.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
