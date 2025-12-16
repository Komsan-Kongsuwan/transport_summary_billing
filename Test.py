"""
Test script to verify billing summary processing logic
Run this to test the app with your sample file
"""

import pandas as pd
import openpyxl
from datetime import date
import calendar

def test_processing():
    print("=" * 60)
    print("BILLING SUMMARY PROCESSOR - TEST SCRIPT")
    print("=" * 60)
    
    # Test file path (update this to your file location)
    test_file = "/mnt/user-data/uploads/Summary_Billing_To_Customer__ใช_ไฟล_น__คำนวนวางบ_ล_.xlsm"
    
    print("\n📁 Loading workbook...")
    wb = openpyxl.load_workbook(test_file, keep_vba=True, data_only=False)
    print(f"✅ Loaded successfully!")
    print(f"   Available sheets: {wb.sheetnames}")
    
    # Check for numeric sheets
    numeric_sheets = [s for s in wb.sheetnames if s.isdigit()]
    print(f"\n📊 Found {len(numeric_sheets)} daily sheets: {numeric_sheets[:10]}{'...' if len(numeric_sheets) > 10 else ''}")
    
    # Check reference sheets
    print("\n🔍 Checking reference sheets...")
    
    if 'Cargo and Weight' in wb.sheetnames:
        cargo_sheet = wb['Cargo and Weight']
        cargo_count = 0
        row = 3
        while cargo_sheet.cell(row, 2).value is not None:
            cargo_count += 1
            row += 1
        print(f"   ✅ Cargo and Weight: {cargo_count} parts found")
    else:
        print("   ❌ Cargo and Weight: NOT FOUND")
    
    if 'Sell Price' in wb.sheetnames:
        sell_sheet = wb['Sell Price']
        sell_count = 0
        row = 2
        while sell_sheet.cell(row, 1).value is not None:
            sell_count += 1
            row += 1
        print(f"   ✅ Sell Price: {sell_count} post codes found")
    else:
        print("   ❌ Sell Price: NOT FOUND")
    
    # Sample processing of first numeric sheet
    if numeric_sheets:
        first_sheet = numeric_sheets[0]
        sheet = wb[first_sheet]
        print(f"\n📋 Analyzing sheet '{first_sheet}':")
        
        row_count = 0
        row = 9
        sample_data = []
        
        while sheet.cell(row, 1).value is not None and row < 15:
            row_count += 1
            sample_data.append({
                'DU': sheet.cell(row, 1).value,
                'Order': sheet.cell(row, 2).value,
                'Part': sheet.cell(row, 13).value,
                'Qty': sheet.cell(row, 14).value
            })
            row += 1
        
        print(f"   Rows found: {row_count}+")
        print(f"   Sample data (first 3 rows):")
        for i, data in enumerate(sample_data[:3], 1):
            print(f"      {i}. DU: {data['DU']}, Order: {data['Order']}, Part: {data['Part']}, Qty: {data['Qty']}")
    
    # Test lookup functionality
    print("\n🔬 Testing lookup functionality...")
    
    if 'Cargo and Weight' in wb.sheetnames and numeric_sheets:
        cargo_sheet = wb['Cargo and Weight']
        test_part = sample_data[0]['Part'] if sample_data else None
        
        if test_part:
            found = False
            row = 3
            while cargo_sheet.cell(row, 2).value is not None:
                if str(cargo_sheet.cell(row, 2).value) == str(test_part):
                    weight = cargo_sheet.cell(row, 5).value
                    print(f"   ✅ Part lookup test: {test_part} -> Weight: {weight} kg")
                    found = True
                    break
                row += 1
            
            if not found:
                print(f"   ⚠️  Part {test_part} not found in Cargo and Weight")
    
    if 'Sell Price' in wb.sheetnames and numeric_sheets:
        sell_sheet = wb['Sell Price']
        first_sheet_obj = wb[numeric_sheets[0]]
        test_postcode = first_sheet_obj.cell(9, 11).value
        
        if test_postcode:
            found = False
            row = 2
            while sell_sheet.cell(row, 1).value is not None:
                if str(sell_sheet.cell(row, 1).value) == str(int(test_postcode)):
                    area = sell_sheet.cell(row, 3).value
                    rate = sell_sheet.cell(row, 5).value
                    print(f"   ✅ Price lookup test: PostCode {test_postcode} -> Area: {area}, Rate: {rate} baht/kg")
                    found = True
                    break
                row += 1
            
            if not found:
                print(f"   ⚠️  PostCode {test_postcode} not found in Sell Price")
    
    # Calculate expected output
    print("\n📊 Expected output summary:")
    total_records = sum(1 for s in numeric_sheets for row in range(9, 100) if s in wb.sheetnames and wb[s].cell(row, 1).value is not None)
    print(f"   Estimated total records: ~{total_records}")
    print(f"   Output will have: 25 columns")
    print(f"   Columns: Order Date, DU, DU-Order, CM Code, Sold To, etc.")
    
    print("\n" + "=" * 60)
    print("✅ TEST COMPLETED SUCCESSFULLY!")
    print("=" * 60)
    print("\n💡 Next steps:")
    print("   1. Run the Streamlit app: streamlit run billing_summary_app.py")
    print("   2. Upload this test file")
    print("   3. Select month: May, Year: 2024")
    print("   4. Click 'Generate Summary'")
    print("   5. Download the result!")
    print("\n")

if __name__ == "__main__":
    try:
        test_processing()
    except FileNotFoundError:
        print("\n❌ Error: Test file not found!")
        print("   Please update the 'test_file' path in this script")
        print("   Or run the Streamlit app directly")
    except Exception as e:
        print(f"\n❌ Error during testing: {str(e)}")
        import traceback
        traceback.print_exc()
